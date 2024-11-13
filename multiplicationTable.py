#! /usr/bin/env python3
#multiplicationTable.py

import sys, openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

number = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.active

bold_font = Font(bold=True)

for i in range(1, number + 1):
    h_row = 'A' + str(i + 1)
    sheet[h_row].value = i
    sheet[h_row].font = bold_font

    h_col = get_column_letter(i + 1) + '1'
    sheet[h_col] = i
    sheet[h_col].font = bold_font

for row in range(1, number + 1):
    for col in range(1, number + 1):
        sheet.cell(row=row + 1, column=col+1).value = row * col
            
wb.save('m.xlsx')
