#! /usr/bin/env python3
#blankRowInserter.py
#python blankRowInserter.py 3 2 myProduce.xlsx

import sys,openpyxl
from openpyxl.utils import get_column_letter

start_row = int(sys.argv[1])
num_rows = int(sys.argv[2])
filename = sys.argv[3]

wb = openpyxl.load_workbook(filename)
sheet = wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for row in range(1, start_row):
    for col in range(1, sheet.max_column + 1):
        new_sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value
        
for row in range(start_row, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        new_sheet.cell(row=row+num_rows, column=col).value = sheet.cell(row=row, column=col).value

new_filename = '_add_black' + filename
new_wb.save(new_filename)
